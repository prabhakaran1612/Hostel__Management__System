import os
import sys
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Add workspace root to Python path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import create_app
from database import fetch_all, fetch_one
from auth_utils import check_password

app = create_app()

def generate_report():
    print("Generating Student Credentials Report...")
    
    with app.app_context():
        # Fetch all female students in Block C
        # Note: We fetch from the students table directly to get their hashes
        students = fetch_all("""
            SELECT name, register_number AS reg_number, department, year, room_no, email, password, gender 
            FROM students 
            WHERE room_no LIKE 'C%' AND gender = 'Female'
            ORDER BY room_no, name
        """)
        
        if not students:
            print("No female student records found in Block C!")
            return
            
        print(f"Retrieved {len(students)} student records.")
        
        # 1. Verification checks
        emails = set()
        failures = 0
        verified_data = []
        
        # Room C406 specific verification tracker
        c406_verified = []
        
        for s in students:
            email = s['email']
            name = s['name']
            
            # Check unique email
            if email in emails:
                print(f"Warning: Duplicate email found: {email}")
            emails.add(email)
            
            # Verify login credentials using project's logic
            # Since we know password used was 'password123'
            plain_password = 'password123'
            is_valid = check_password(plain_password, s['password'])
            
            if not is_valid:
                print(f"Auth Verification FAILED for student: {name} ({email})")
                failures += 1
            else:
                if s['room_no'] == 'C406':
                    c406_verified.append(name)
            
            verified_data.append({
                'Name': name,
                'Register Number': s['reg_number'],
                'Department': s['department'],
                'Year': s['year'],
                'Room Number': s['room_no'],
                'Email': email,
                'Password': plain_password,
                'Gender': s['gender']
            })
            
        print(f"Authentication verification completed: {len(students) - failures} passed, {failures} failed.")
        
        # Specifically verify C406 names list
        expected_c406 = {"Rashvi", "Aaslin", "Charu Harinee", "Yamini"}
        actual_c406 = set(c406_verified)
        missing_c406 = expected_c406 - actual_c406
        if not missing_c406:
            print("OK - C406 specific student verification (Rashvi, Aaslin, Charu Harinee, Yamini) passed successfully.")
        else:
            print(f"Error: Missing or failed C406 students: {missing_c406}")
            
        # 2. Export to CSV
        csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'student_credentials.csv')
        headers = ['Name', 'Register Number', 'Department', 'Year', 'Room Number', 'Email', 'Password', 'Gender']
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(verified_data)
            
        print(f"OK - Exported CSV report to: {csv_path}")

        # 3. Export to Excel (.xlsx) using openpyxl
        xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'student_credentials.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Credentials"
        
        # Styling parameters
        header_font = Font(name='Sora', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Indigo fill
        cell_font = Font(name='Sora', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Write headers
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align if col_idx in [2, 3, 4, 5, 8] else left_align
            
        # Write data rows
        for row_idx, data in enumerate(verified_data, start=2):
            row_values = [data[h] for h in headers]
            ws.append(row_values)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.alignment = center_align if col_idx in [2, 3, 4, 5, 8] else left_align
                
        # Set column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(xlsx_path)
        print(f"OK - Exported Excel report to: {xlsx_path}")
        
        # 4. Generate Summaries
        # Room-wise count
        room_counts = {}
        # Department-wise count
        dept_counts = {}
        
        for data in verified_data:
            r = data['Room Number']
            room_counts[r] = room_counts.get(r, 0) + 1
            
            d = data['Department']
            dept_counts[d] = dept_counts.get(d, 0) + 1
            
        # 5. Output Summary and First 20 records
        print("\n" + "="*50)
        print(" credentials generation report ".upper().center(50, "="))
        print("="*50)
        print(f"Total students created: {len(verified_data)}")
        
        print("\nStudents Per Room:")
        for r_no in sorted(room_counts.keys()):
            print(f"- Room {r_no}: {room_counts[r_no]}")
            
        print("\nStudents Per Department:")
        for dept in sorted(dept_counts.keys()):
            print(f"- {dept}: {dept_counts[dept]}")
            
        print("\n" + "-"*50)
        print(" first 20 student credentials ".upper().center(50, "-"))
        print("-"*50)
        for idx, s in enumerate(verified_data[:20], start=1):
            print(f"{idx:02d}. Name: {s['Name']:<15} | Reg: {s['Register Number']:<10} | Room: {s['Room Number']:<6} | Email: {s['Email']:<30} | Plain Pwd: {s['Password']}")
        print("="*50 + "\n")

if __name__ == '__main__':
    generate_report()
